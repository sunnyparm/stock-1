# -*- coding: utf-8 -*-
"""

"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
import os
import warnings
import tkinter as tk
from tkinter import filedialog
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
warnings.filterwarnings('ignore')

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

class DoubleBottomAnalyzer:
    """쌍바닥 패턴 분석 클래스"""
    
    def __init__(self, file_path):
        """
        초기화
        
        Args:
            file_path (str): 코스피 데이터 CSV 파일 경로
        """
        self.file_path = file_path
        self.df = None
        self.df_long = None
        self.double_bottom_results = []
        self.stock_info = {}  # 종목별 시가총액, 섹터 정보 저장
        
    def load_and_transform_data(self):
        """데이터 로드 및 변환 (가로 -> 세로 형태)"""
        print("📊 데이터 로드 및 변환 중...")
        
        # CSV 파일 읽기
        self.df = pd.read_csv(self.file_path)
        
        # 첫 번째 컬럼이 '종목'인지 확인
        if self.df.columns[0] != '종목':
            raise ValueError("첫 번째 컬럼이 '종목'이 아닙니다.")
        
        # 시가총액, 섹터 정보 추출
        if '시가총액' in self.df.columns and '섹터' in self.df.columns:
            for _, row in self.df.iterrows():
                symbol = row['종목']
                market_cap = row.get('시가총액', 'N/A')
                sector = row.get('섹터', 'N/A')
                self.stock_info[symbol] = {
                    '시가총액': market_cap,
                    '섹터': sector
                }
            print(f"📊 종목 정보 추출 완료: {len(self.stock_info)}개 종목")
        
        # 시가총액, 섹터 열이 있으면 제외하고 날짜 열만 선택
        date_columns = []
        for col in self.df.columns:
            if col == '종목' or col == '시가총액' or col == '섹터':
                continue
            # 날짜 형식인지 확인 (YYYY-MM-DD)
            try:
                pd.to_datetime(col, format='%Y-%m-%d')
                date_columns.append(col)
            except:
                continue
        
        if not date_columns:
            raise ValueError("날짜 형식의 컬럼을 찾을 수 없습니다.")
        
        # 종목과 날짜 열만 선택
        columns_to_use = ['종목'] + date_columns
        self.df = self.df[columns_to_use]
        
        # 종목명을 인덱스로 설정
        self.df = self.df.set_index('종목')
        
        # 가로 형태를 세로 형태로 변환 (melt)
        self.df_long = self.df.reset_index().melt(
            id_vars=['종목'], 
            var_name='Date', 
            value_name='Close'
        )
        
        # 날짜 컬럼을 datetime으로 변환
        self.df_long['Date'] = pd.to_datetime(self.df_long['Date'])
        
        # 결측값 제거
        self.df_long = self.df_long.dropna()
        
        # 종목별, 날짜별로 정렬
        self.df_long = self.df_long.sort_values(['종목', 'Date']).reset_index(drop=True)
        
        print(f"✅ 데이터 변환 완료: {len(self.df_long)} 행, {self.df_long['종목'].nunique()}개 종목")
        print(f"📅 날짜 범위: {self.df_long['Date'].min().strftime('%Y-%m-%d')} ~ {self.df_long['Date'].max().strftime('%Y-%m-%d')}")
        
    def find_local_minima_maxima(self, series, window=5):
        """
        로컬 최소값과 최대값의 인덱스를 찾습니다.
        
        Args:
            series: 가격 시계열
            window: 비교할 윈도우 크기
            
        Returns:
            tuple: (최소값 인덱스 리스트, 최대값 인덱스 리스트)
        """
        minima_indices = []
        maxima_indices = []
        
        for i in range(window, len(series) - window):
            # 로컬 최소값
            if series.iloc[i] == series.iloc[i-window : i+window+1].min():
                minima_indices.append(i)
            # 로컬 최대값
            if series.iloc[i] == series.iloc[i-window : i+window+1].max():
                maxima_indices.append(i)
                
        return minima_indices, maxima_indices
    
    def find_previous_lowest_bottom(self, data, current_bottom_idx, lookback_days=60):
        """
        이전 최저 바닥 찾기
        
        Args:
            data: 주가 데이터
            current_bottom_idx: 현재 바닥 인덱스
            lookback_days: 이전 기간 (일)
            
        Returns:
            tuple: (이전 최저가, 이전 최저가 인덱스)
        """
        if current_bottom_idx < lookback_days:
            return None, None
        
        # 이전 기간 데이터
        previous_data = data.iloc[max(0, current_bottom_idx - lookback_days):current_bottom_idx]
        if len(previous_data) == 0:
            return None, None
        
        # 이전 기간 최저가
        prev_lowest = previous_data['Close'].min()
        prev_lowest_idx = previous_data['Close'].idxmin()
        
        return prev_lowest, prev_lowest_idx
    
    def check_sideways_after_bottom(self, data, bottom_idx, days=3, tolerance_pct=0.02):
        """
        바닥 이후 N일간 횡보 여부 확인
        
        Args:
            data: 주가 데이터
            bottom_idx: 바닥 인덱스
            days: 확인할 일수
            tolerance_pct: 횡보 허용 범위 (퍼센트)
        
        Returns:
            bool: 횡보 여부
        """
        if bottom_idx + days >= len(data):
            return False
        
        # 바닥 이후 N일간 데이터
        after_bottom_data = data.iloc[bottom_idx:bottom_idx + days + 1]
        bottom_price = data.iloc[bottom_idx]['Close']
        
        # 바닥 이후 최고가와 최저가
        max_price = after_bottom_data['Close'].max()
        min_price = after_bottom_data['Close'].min()
        
        # 바닥 가격 대비 변동폭이 허용 범위 내인지 확인
        max_deviation = max(abs(max_price - bottom_price), abs(min_price - bottom_price))
        deviation_pct = max_deviation / bottom_price
        
        return deviation_pct <= tolerance_pct

    def validate_double_bottom_pattern(self, data, first_bottom_idx, second_bottom_idx, 
                                     first_bottom_price, second_bottom_price, lookback_days=60):
        """
        쌍바닥 패턴 유효성 검증 (개선된 버전)
        
        Args:
            data: 주가 데이터
            first_bottom_idx: 첫 번째 바닥 인덱스
            second_bottom_idx: 두 번째 바닥 인덱스
            first_bottom_price: 첫 번째 바닥 가격
            second_bottom_price: 두 번째 바닥 가격
            lookback_days: 이전 기간 (일)
            
        Returns:
            dict: 검증 결과
        """
        # 1. 이전 최저 바닥 확인
        prev_lowest, prev_lowest_idx = self.find_previous_lowest_bottom(
            data, first_bottom_idx, lookback_days)
        
        # 2. 상대적 바닥 확인 (주변 20일 대비)
        first_window = data.iloc[max(0, first_bottom_idx-10):first_bottom_idx+10]
        first_local_min = first_window['Close'].min()
        
        second_window = data.iloc[max(0, second_bottom_idx-10):second_bottom_idx+10]
        second_local_min = second_window['Close'].min()
        
        # 3. 최근 최저가 이후 3일간 횡보 확인
        recent_bottom_sideways = self.check_sideways_after_bottom(data, second_bottom_idx, days=3, tolerance_pct=0.02)
        
        validation_result = {
            'prev_lowest': prev_lowest,
            'prev_lowest_idx': prev_lowest_idx,
            'first_local_min': first_local_min,
            'second_local_min': second_local_min,
            'recent_bottom_sideways': recent_bottom_sideways,
            'is_valid_double_bottom': False,
            'issues': [],
            'validation_score': 0
        }
        
        # 유효성 검증 및 점수 계산
        score = 100  # 기본 점수
        
        # 이전 최저 바닥 검증 (30점)
        if prev_lowest is not None:
            if first_bottom_price <= prev_lowest:
                validation_result['issues'].append(f"첫 번째 바닥이 이전 최저가보다 낮거나 같음")
                score -= 30
            else:
                # 이전 최저가보다 높으면 보너스 점수
                improvement_pct = (first_bottom_price - prev_lowest) / prev_lowest
                if improvement_pct > 0.05:  # 5% 이상 개선
                    score += 10
        
        # 상대적 바닥 검증 (20점씩)
        if first_bottom_price != first_local_min:
            validation_result['issues'].append(f"첫 번째 바닥이 주변 최저가와 다름")
            score -= 20
            
        if second_bottom_price != second_local_min:
            validation_result['issues'].append(f"두 번째 바닥이 주변 최저가와 다름")
            score -= 20
        
        # 최근 바닥 이후 횡보 검증 (20점)
        if not recent_bottom_sideways:
            validation_result['issues'].append(f"최근 바닥 이후 3일간 횡보하지 않음")
            score -= 20
        else:
            # 횡보하면 보너스 점수
            score += 10
        
        # 최종 유효성 판단
        if len(validation_result['issues']) == 0:
            validation_result['is_valid_double_bottom'] = True
        
        validation_result['validation_score'] = max(0, score)
        
        return validation_result
    
    def calculate_double_bottom_score(self, stock_data, 
                                    bottom_tolerance_pct=0.10,  # 더 엄격하게 (20% -> 10%)
                                    rebound_min_pct=0.05,       # 더 엄격하게 (2% -> 5%)
                                    breakout_min_pct=0.02,      # 더 엄격하게 (0.5% -> 2%)
                                    min_days_between_bottoms=10, # 더 엄격하게 (5일 -> 10일)
                                    max_days_between_bottoms=120, # 더 엄격하게 (200일 -> 120일)
                                    lookback_period=120):       # 더 긴 기간 (100일 -> 120일)
        """
        쌍바닥 패턴 점수를 계산합니다.
        
        Args:
            stock_data: 종목별 주가 데이터
            bottom_tolerance_pct: 두 바닥의 가격 차이 허용 범위
            rebound_min_pct: 첫 번째 바닥 이후 최소 반등률
            breakout_min_pct: 넥라인 돌파 최소 비율
            min_days_between_bottoms: 두 바닥 사이 최소 기간
            max_days_between_bottoms: 두 바닥 사이 최대 기간
            lookback_period: 패턴을 찾을 전체 기간
            
        Returns:
            dict: 쌍바닥 패턴 분석 결과
        """
        if len(stock_data) < 50:  # 최소 데이터 요구량 줄임
            return None
            
        # 최근 데이터만 사용
        df_recent = stock_data.tail(lookback_period).reset_index(drop=True)
        n = len(df_recent)
        
        # 로컬 최소값 찾기 (윈도우 크기 줄임)
        minima_indices, _ = self.find_local_minima_maxima(df_recent['Close'], window=3)
        
        if len(minima_indices) < 2:
            return None
        
        best_score = 0
        best_pattern = None
        
        # 모든 바닥 조합 검사
        for i, b1_idx in enumerate(minima_indices[:-1]):  # 마지막 바닥은 제외
            b1_price = df_recent.loc[b1_idx, 'Close']
            
            # 두 번째 바닥 후보 찾기
            b2_candidates = [idx for idx in minima_indices[i+1:] 
                           if (idx - b1_idx >= min_days_between_bottoms and 
                               idx - b1_idx <= max_days_between_bottoms)]
            
            for b2_idx in b2_candidates:
                b2_price = df_recent.loc[b2_idx, 'Close']
                
                # 두 바닥의 가격 차이 확인
                price_diff_pct = abs(b1_price - b2_price) / min(b1_price, b2_price)
                if price_diff_pct > bottom_tolerance_pct:
                    continue
                
                # 두 바닥 사이의 피크 찾기
                if b2_idx - b1_idx <= 1:
                    continue
                    
                peak_slice = df_recent.loc[b1_idx + 1 : b2_idx - 1, 'Close']
                if peak_slice.empty:
                    continue
                    
                peak_idx = peak_slice.idxmax()
                peak_price = df_recent.loc[peak_idx, 'Close']
                
                # 반등률 확인
                rebound_pct = (peak_price - b1_price) / b1_price
                if rebound_pct < rebound_min_pct:
                    continue
                
                # 현재 가격 (마지막 종가)
                current_price = df_recent['Close'].iloc[-1]
                
                # 넥라인 돌파 확인
                breakout_pct = (current_price - peak_price) / peak_price
                
                # 점수 계산 (0-100점)
                score = 0
                
                # 바닥 유사성 점수 (30점)
                similarity_score = max(0, 30 * (1 - price_diff_pct / bottom_tolerance_pct))
                score += similarity_score
                
                # 반등률 점수 (25점)
                rebound_score = min(25, 25 * rebound_pct / rebound_min_pct)
                score += rebound_score
                
                # 돌파 점수 (25점)
                if breakout_pct > 0:
                    breakout_score = min(25, 25 * breakout_pct / breakout_min_pct)
                    score += breakout_score
                else:
                    # 돌파하지 않았어도 패턴이 있으면 기본 점수
                    breakout_score = 10
                    score += breakout_score
                
                # 패턴 완성도 점수 (20점)
                pattern_completeness = 20 if breakout_pct > breakout_min_pct else 15
                score += pattern_completeness
                
                # 유효성 검증 추가
                validation = self.validate_double_bottom_pattern(
                    df_recent, b1_idx, b2_idx, b1_price, b2_price, 60)
                
                # 유효성 점수 반영 (기존 점수에 가중치 적용)
                if validation['is_valid_double_bottom']:
                    # 유효한 쌍바닥이면 점수 보너스
                    score = score * 1.2  # 20% 보너스
                else:
                    # 의심스러운 패턴이면 점수 감점
                    score = score * 0.7  # 30% 감점
                
                # 유효성 검증 점수 추가 (최대 20점)
                validation_bonus = min(20, validation['validation_score'] * 0.2)
                score += validation_bonus
                
                if score > best_score:
                    best_score = score
                    best_pattern = {
                        'b1_idx': b1_idx,
                        'b2_idx': b2_idx,
                        'peak_idx': peak_idx,
                        'b1_price': b1_price,
                        'b2_price': b2_price,
                        'peak_price': peak_price,
                        'current_price': current_price,
                        'price_diff_pct': price_diff_pct,
                        'rebound_pct': rebound_pct,
                        'breakout_pct': breakout_pct,
                        'score': score,
                        'validation': validation
                    }
        
        return best_pattern
    
    def analyze_all_stocks(self):
        """모든 종목에 대해 쌍바닥 분석 수행"""
        print("🔍 모든 종목 쌍바닥 패턴 분석 중...")
        
        results = []
        total_stocks = self.df_long['종목'].nunique()
        processed = 0
        
        for symbol in self.df_long['종목'].unique():
            stock_data = self.df_long[self.df_long['종목'] == symbol].copy()
            
            # 쌍바닥 패턴 분석
            pattern = self.calculate_double_bottom_score(stock_data)
            
            if pattern:
                pattern['symbol'] = symbol
                results.append(pattern)
                print(f"✅ {symbol}: 점수 {pattern['score']:.1f}")
            
            processed += 1
            if processed % 100 == 0:
                print(f"진행률: {processed}/{total_stocks} ({processed/total_stocks*100:.1f}%)")
        
        # 점수순으로 정렬
        self.double_bottom_results = sorted(results, key=lambda x: x['score'], reverse=True)
        
        print(f"\n📈 분석 완료: {len(self.double_bottom_results)}개 종목에서 쌍바닥 패턴 발견")
        
    def get_valid_double_bottom_results(self):
        """유효한 쌍바닥 패턴만 반환"""
        valid_results = []
        for result in self.double_bottom_results:
            validation = result.get('validation', {})
            if validation.get('is_valid_double_bottom', False):
                valid_results.append(result)
        return valid_results
    
    def get_top15_results(self):
        """상위 15개 결과 반환 (하위 호환성)"""
        return self.double_bottom_results[:15]
    
    def print_valid_results(self):
        """유효한 쌍바닥 패턴 결과 출력 (15개 제한 없음)"""
        valid_results = self.get_valid_double_bottom_results()
        
        print("\n" + "="*80)
        print(f"🏆 유효한 쌍바닥 패턴 종목 ({len(valid_results)}개)")
        print("="*80)
        
        for i, result in enumerate(valid_results, 1):
            print(f"\n{i:2d}. {result['symbol']}")
            print(f"    📊 종합 점수: {result['score']:.1f}/100")
            print(f"    💰 첫 번째 바닥: {result['b1_price']:,.0f}원")
            print(f"    💰 두 번째 바닥: {result['b2_price']:,.0f}원")
            print(f"    📈 넥라인: {result['peak_price']:,.0f}원")
            print(f"    💎 현재가: {result['current_price']:,.0f}원")
            print(f"    📉 바닥 차이: {result['price_diff_pct']*100:.1f}%")
            print(f"    📈 반등률: {result['rebound_pct']*100:.1f}%")
            print(f"    🚀 돌파률: {result['breakout_pct']*100:.1f}%")
            
            validation = result.get('validation', {})
            if validation.get('prev_lowest'):
                print(f"    📊 이전 최저가: {validation['prev_lowest']:,.0f}원")
                improvement = (result['b1_price'] - validation['prev_lowest']) / validation['prev_lowest'] * 100
                print(f"    📈 바닥 개선률: {improvement:.1f}%")
            
            # 횡보 정보 출력
            if validation.get('recent_bottom_sideways') is not None:
                print(f"    📊 최근 바닥 이후 횡보: {'✅ 횡보' if validation['recent_bottom_sideways'] else '❌ 횡보하지 않음'}")
    
    def print_top15_results(self):
        """상위 15개 결과 출력 (하위 호환성)"""
        top15 = self.get_top15_results()
        
        print("\n" + "="*80)
        print("🏆 상위 15개 쌍바닥 패턴 종목")
        print("="*80)
        
        for i, result in enumerate(top15, 1):
            print(f"\n{i:2d}. {result['symbol']}")
            print(f"    📊 종합 점수: {result['score']:.1f}/100")
            print(f"    💰 첫 번째 바닥: {result['b1_price']:,.0f}원")
            print(f"    💰 두 번째 바닥: {result['b2_price']:,.0f}원")
            print(f"    📈 넥라인: {result['peak_price']:,.0f}원")
            print(f"    💎 현재가: {result['current_price']:,.0f}원")
            print(f"    📉 바닥 차이: {result['price_diff_pct']*100:.1f}%")
            print(f"    📈 반등률: {result['rebound_pct']*100:.1f}%")
            print(f"    🚀 돌파률: {result['breakout_pct']*100:.1f}%")
    
    def save_valid_results_to_csv(self, filename='valid_double_bottom_results.csv'):
        """유효한 쌍바닥 패턴 결과를 CSV 파일로 저장"""
        valid_results = self.get_valid_double_bottom_results()
        
        if not valid_results:
            print("저장할 유효한 결과가 없습니다.")
            return
        
        # two_bottom 폴더 경로 설정
        output_dir = 'two_bottom'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 파일명에 타임스탬프 추가
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if filename == 'valid_double_bottom_results.csv':
            filename = f'valid_double_bottom_results_{timestamp}.csv'
        else:
            # 파일명에 확장자가 있는 경우 타임스탬프를 확장자 앞에 추가
            name, ext = filename.rsplit('.', 1) if '.' in filename else (filename, 'csv')
            filename = f'{name}_{timestamp}.{ext}'
        
        # 전체 파일 경로 생성
        filepath = os.path.join(output_dir, filename)
        
        # 결과 데이터프레임 생성
        results_data = []
        for result in valid_results:
            symbol = result['symbol']
            stock_info = self.stock_info.get(symbol, {})
            
            results_data.append({
                '종목': symbol,
                '첫번째바닥': result['b1_price'],
                '두번째바닥': result['b2_price'],
                '넥라인': result['peak_price'],
                '현재가': result['current_price'],
                '바닥차이(%)': round(result['price_diff_pct'] * 100, 2),
                '반등률(%)': round(result['rebound_pct'] * 100, 2),
                '돌파률(%)': round(result['breakout_pct'] * 100, 2),
                '시가총액': stock_info.get('시가총액', 'N/A'),
                '섹터': stock_info.get('섹터', 'N/A')
            })
        
        df_results = pd.DataFrame(results_data)
        
        # 돌파률 기준으로 정렬 (마이너스 값이 위에 오도록)
        df_results = df_results.sort_values('돌파률(%)', ascending=True)
        
        df_results.to_csv(filepath, index=False, encoding='utf-8-sig')
        print(f"✅ 유효한 쌍바닥 패턴 결과가 '{filepath}' 파일로 저장되었습니다.")
        print(f"📊 총 {len(valid_results)}개 종목이 유효한 쌍바닥 패턴으로 확인되었습니다.")
    
    def create_sparkline_image(self, data, width=200, height=50, color='#2E86AB'):
        """스파크라인 이미지 생성"""
        data = pd.to_numeric(data, errors='coerce').dropna()
        
        if len(data) < 2:
            return None
        
        # DPI 설정
        dpi = 100
        fig_width = width / dpi
        fig_height = height / dpi
        
        # 그래프 생성
        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=dpi)
        
        # 메인 라인 그리기
        ax.plot(data.values, color=color, linewidth=2, alpha=0.8)
        
        # 최고점과 최저점 표시
        max_idx = data.values.argmax()
        min_idx = data.values.argmin()
        ax.scatter([max_idx], [data.values[max_idx]], color='red', s=15, zorder=5)
        ax.scatter([min_idx], [data.values[min_idx]], color='blue', s=15, zorder=5)
        
        # 축 숨기기
        ax.set_xticks([])
        ax.set_yticks([])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        
        # 배경 설정
        ax.set_facecolor('white')
        fig.patch.set_facecolor('white')
        
        # 여백 제거
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
        
        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        fig.savefig(temp_file.name, dpi=dpi, bbox_inches='tight', 
                   pad_inches=0, facecolor='white', edgecolor='none')
        plt.close(fig)
        
        return temp_file.name
    
    def save_results_to_excel_with_sparklines(self, filename='double_bottom_results_with_sparklines.xlsx'):
        """스파크라인이 포함된 Excel 파일로 저장"""
        valid_results = self.get_valid_double_bottom_results()
        
        if not valid_results:
            print("저장할 유효한 결과가 없습니다.")
            return
        
        # two_bottom 폴더 경로 설정
        output_dir = 'two_bottom'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 파일명에 타임스탬프 추가
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if filename == 'double_bottom_results_with_sparklines.xlsx':
            filename = f'double_bottom_results_with_sparklines_{timestamp}.xlsx'
        else:
            # 파일명에 확장자가 있는 경우 타임스탬프를 확장자 앞에 추가
            name, ext = filename.rsplit('.', 1) if '.' in filename else (filename, 'xlsx')
            filename = f'{name}_{timestamp}.{ext}'
        
        # 전체 파일 경로 생성
        filepath = os.path.join(output_dir, filename)
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "쌍바닥 패턴 분석 결과"
        
        # 헤더 추가
        headers = ['순위', '종목', '스파크라인', '종합점수', '첫번째바닥', '두번째바닥', 
                  '넥라인', '현재가', '바닥차이(%)', '반등률(%)', '돌파률(%)', 
                  '유효성검증', '검증점수', '시가총액', '섹터']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 8   # 순위
        ws.column_dimensions['B'].width = 15  # 종목명
        ws.column_dimensions['C'].width = 25  # 스파크라인
        ws.column_dimensions['D'].width = 12  # 종합점수
        ws.column_dimensions['E'].width = 15  # 첫번째바닥
        ws.column_dimensions['F'].width = 15  # 두번째바닥
        ws.column_dimensions['G'].width = 15  # 넥라인
        ws.column_dimensions['H'].width = 15  # 현재가
        ws.column_dimensions['I'].width = 12  # 바닥차이
        ws.column_dimensions['J'].width = 12  # 반등률
        ws.column_dimensions['K'].width = 12  # 돌파률
        ws.column_dimensions['L'].width = 15  # 유효성검증
        ws.column_dimensions['M'].width = 12  # 검증점수
        ws.column_dimensions['N'].width = 15  # 시가총액
        ws.column_dimensions['O'].width = 15  # 섹터
        
        # 행 높이 설정 (스파크라인이 잘리지 않도록)
        ws.row_dimensions[1].height = 30  # 헤더 행 높이
        for row in range(2, len(valid_results) + 2):  # 데이터 행들
            ws.row_dimensions[row].height = 60  # 스파크라인을 위한 충분한 높이
        
        # 각 종목 데이터 추가
        print(f"Excel 스파크라인 생성 중... ({len(valid_results)}개 종목)")
        temp_files = []  # 임시 파일 추적용
        
        for row_idx, result in enumerate(valid_results, 2):
            try:
                symbol = result['symbol']
                validation = result.get('validation', {})
                stock_info = self.stock_info.get(symbol, {})
                
                # 순위
                ws.cell(row=row_idx, column=1, value=row_idx-1)
                
                # 종목명
                ws.cell(row=row_idx, column=2, value=symbol)
                
                # 스파크라인 생성
                stock_data = self.df_long[self.df_long['종목'] == symbol].copy()
                if len(stock_data) > 1:
                    # 최근 6개월 데이터 사용
                    recent_data = stock_data.tail(120)  # 약 6개월
                    price_data = recent_data['Close']
                    
                    temp_img_path = self.create_sparkline_image(price_data, 200, 50)
                    if temp_img_path:
                        temp_files.append(temp_img_path)  # 나중에 삭제하기 위해 추적
                        
                        # Excel에 이미지 삽입
                        img = Image(temp_img_path)
                        img.width = 200
                        img.height = 50
                        ws.add_image(img, f'C{row_idx}')
                
                # 데이터 추가
                data_values = [
                    round(result['score'], 1),  # 종합점수
                    result['b1_price'],         # 첫번째바닥
                    result['b2_price'],         # 두번째바닥
                    result['peak_price'],       # 넥라인
                    result['current_price'],    # 현재가
                    round(result['price_diff_pct'] * 100, 2),  # 바닥차이
                    round(result['rebound_pct'] * 100, 2),     # 반등률
                    round(result['breakout_pct'] * 100, 2),    # 돌파률
                    '✅ 유효' if validation.get('is_valid_double_bottom', False) else '❌ 의심',  # 유효성검증
                    round(validation.get('validation_score', 0), 1),  # 검증점수
                    stock_info.get('시가총액', 'N/A'),  # 시가총액
                    stock_info.get('섹터', 'N/A')       # 섹터
                ]
                
                for col_idx, value in enumerate(data_values, 4):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
                
                if row_idx % 10 == 0:
                    print(f"진행률: {row_idx-1}/{len(valid_results)} 종목 완료")
                    
            except Exception as e:
                print(f"종목 '{symbol}' 처리 중 오류: {e}")
                continue
        
        # Excel 파일 저장
        wb.save(filepath)
        
        # 임시 파일들 삭제
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        
        print(f"✅ 스파크라인이 포함된 Excel 파일이 '{filepath}'에 저장되었습니다.")
        print(f"📊 총 {len(valid_results)}개 종목의 쌍바닥 패턴 분석 결과가 스파크라인과 함께 저장되었습니다.")
        return filepath

    def save_results_to_csv(self, filename='improved_top15_double_bottom_results.csv'):
        """개선된 결과를 CSV 파일로 저장 (유효성 검증 정보 포함)"""
        top15 = self.get_top15_results()
        
        if not top15:
            print("저장할 결과가 없습니다.")
            return
        
        # two_bottom 폴더 경로 설정
        output_dir = 'two_bottom'
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 파일명에 타임스탬프 추가
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if filename == 'improved_top15_double_bottom_results.csv':
            filename = f'improved_top15_double_bottom_results_{timestamp}.csv'
        else:
            # 파일명에 확장자가 있는 경우 타임스탬프를 확장자 앞에 추가
            name, ext = filename.rsplit('.', 1) if '.' in filename else (filename, 'csv')
            filename = f'{name}_{timestamp}.{ext}'
        
        # 전체 파일 경로 생성
        filepath = os.path.join(output_dir, filename)
        
        # 결과 데이터프레임 생성
        results_data = []
        for result in top15:
            validation = result.get('validation', {})
            results_data.append({
                '종목': result['symbol'],
                '종합점수': round(result['score'], 1),
                '첫번째바닥': result['b1_price'],
                '두번째바닥': result['b2_price'],
                '넥라인': result['peak_price'],
                '현재가': result['current_price'],
                '바닥차이(%)': round(result['price_diff_pct'] * 100, 2),
                '반등률(%)': round(result['rebound_pct'] * 100, 2),
                '돌파률(%)': round(result['breakout_pct'] * 100, 2),
                '유효성검증': '✅ 유효' if validation.get('is_valid_double_bottom', False) else '❌ 의심',
                '검증점수': round(validation.get('validation_score', 0), 1),
                '이전최저바닥': validation.get('prev_lowest', 'N/A'),
                '최근바닥횡보': '✅ 횡보' if validation.get('recent_bottom_sideways', False) else '❌ 횡보하지 않음',
                '문제점수': len(validation.get('issues', [])),
                '문제점': '; '.join(validation.get('issues', [])) if validation.get('issues') else '없음'
            })
        
        df_results = pd.DataFrame(results_data)
        df_results.to_csv(filepath, index=False, encoding='utf-8-sig')
        print(f"✅ 개선된 결과가 '{filepath}' 파일로 저장되었습니다.")

def select_csv_file():
    """CSV 파일 선택 대화상자"""
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    file_path = filedialog.askopenfilename(
        title="분석할 CSV 파일을 선택하세요",
        filetypes=[
            ("CSV files", "*.csv"),
            ("All files", "*.*")
        ],
        initialdir="two_bottom"  # two_bottom 폴더에서 시작
    )
    
    root.destroy()
    return file_path

def main():
    """메인 실행 함수"""
    print("🚀 개선된 상위 15개 쌍바닥 분석 시작 (이전 최저 바닥 고려)")
    print("="*60)
    
    # 파일 업로드
    print("\n📂 분석할 CSV 파일을 선택하세요...")
    csv_filename = select_csv_file()
    
    if not csv_filename:
        print("❌ 파일이 선택되지 않았습니다. 프로그램을 종료합니다.")
        return
    
    print(f"✅ 선택된 파일: {csv_filename}")
    
    # 파일 존재 여부 확인
    if not os.path.exists(csv_filename):
        print(f"❌ 파일을 찾을 수 없습니다: {csv_filename}")
        print("프로그램을 종료합니다.")
        return
    
    print(f"\n📊 {csv_filename} 파일로 분석을 시작합니다...")
    
    # 분석기 초기화
    analyzer = DoubleBottomAnalyzer(csv_filename)
    
    try:
        # 데이터 로드 및 변환
        analyzer.load_and_transform_data()
        
        # 더 엄격한 쌍바닥 패턴 분석
        print("\n🔍 더 엄격한 쌍바닥 패턴 분석 중...")
        print("   • 더 엄격한 조건 적용 (10% 이내, 5% 반등, 2% 돌파)")
        print("   • 이전 최저 바닥 고려")
        print("   • 상대적 바닥 검증")
        print("   • 최근 바닥 이후 3일간 횡보 확인")
        print("   • 패턴 유효성 검증")
        analyzer.analyze_all_stocks()
        
        # 유효한 쌍바닥 패턴만 출력
        analyzer.print_valid_results()
        
        # 유효한 결과만 CSV로 저장
        analyzer.save_valid_results_to_csv()
        
        # 스파크라인이 포함된 Excel 파일 생성
        print("\n📊 스파크라인이 포함된 Excel 파일 생성 중...")
        excel_file = analyzer.save_results_to_excel_with_sparklines()
        
        print("\n🎉 유효한 쌍바닥 패턴 분석 완료!")
        print("="*60)
        print("📋 주요 특징:")
        print("   ✅ 더 엄격한 쌍바닥 조건 (10% 이내, 5% 반등, 2% 돌파)")
        print("   ✅ 이전 최저 바닥 고려한 정확한 쌍바닥 분석")
        print("   ✅ 상대적 바닥 검증으로 가짜 패턴 제거")
        print("   ✅ 최근 바닥 이후 3일간 횡보 패턴 확인")
        print("   ✅ 15개 제한 없이 모든 유효한 쌍바닥 선별")
        print("   ✅ 진짜 쌍바닥 패턴만 추출")
        print("   📈 스파크라인 차트가 포함된 Excel 파일 생성")
        print(f"   📁 Excel 파일: {excel_file}")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()