import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
import os
import platform
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import tempfile
import warnings
warnings.filterwarnings('ignore')

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

def detect_late_trough(series: pd.Series,
                       trough_threshold: float = -0.1,
                       late_trough_cutoff: float = 0.9) -> bool:
    """
    공통 패턴: '후반부 저점' 여부 감지 (반등률 조건 제거)
    
    Parameters
    ----------
    series : pd.Series
        정규화된 종가 시계열 (첫날=100 기준).
    trough_threshold : float
        시작 대비 저점 하락률 조건 (예: -0.1 → -10% 이상 하락).
    late_trough_cutoff : float
        저점이 발생하는 시점 기준 (0~1 비율, 예: 0.9 → 전체의 90% 이후여야 후반부 저점으로 인정).
    
    Returns
    -------
    bool
        조건을 만족하면 True, 아니면 False
    """
    s = series.dropna()
    if len(s) < 10:
        return False
    
    n = len(s)
    start_val = s.iloc[0]
    argmin = int(np.argmin(s.values))
    trough_val = s.iloc[argmin]
    
    # 1) 시작 대비 충분한 하락
    drop_from_start = trough_val / start_val - 1.0
    cond_drop = drop_from_start <= trough_threshold
    
    # 2) 저점 시점이 후반부에 존재 (90% 이후)
    pos_min = argmin / (n - 1)
    cond_late_trough = pos_min >= late_trough_cutoff
    
    return cond_drop and cond_late_trough

def create_sparkline_image(data, width=180, height=40, color='#2E86AB'):
    """스파크라인 이미지 생성"""
    data = pd.to_numeric(data, errors='coerce').dropna()
    
    if len(data) < 2:
        return None
    
    # DPI 설정
    dpi = 100
    fig_width = width / dpi
    fig_height = height / dpi
    
    # 그래프 생성
    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=dpi)
    
    # 메인 라인 그리기
    ax.plot(data.values, color=color, linewidth=2, alpha=0.8)
    
    # 최고점과 최저점 표시
    max_idx = data.values.argmax()
    min_idx = data.values.argmin()
    ax.scatter([max_idx], [data.values[max_idx]], color='red', s=15, zorder=5)
    ax.scatter([min_idx], [data.values[min_idx]], color='blue', s=15, zorder=5)
    
    # 축 숨기기
    ax.set_xticks([])
    ax.set_yticks([])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    
    # 배경 설정
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')
    
    # 여백 제거
    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    
    # 임시 파일로 저장
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    fig.savefig(temp_file.name, dpi=dpi, bbox_inches='tight', 
               pad_inches=0, facecolor='white', edgecolor='none')
    plt.close(fig)
    
    return temp_file.name

# =========================
# 🔧 사용자 설정
# =========================
TROUGH_THRESHOLD = -0.05   # 저점 하락률 (-5% - 완화)
LATE_TROUGH_CUTOFF = 0.8   # 후반부 저점 기준 (80% - 완화)

# =========================
# 파일 업로드
# =========================
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(
    title="코스피 6개월 종가 파일 업로드",
    filetypes=[("CSV files", "*.csv")]
)

df = pd.read_csv(file_path)

# =========================
# 날짜 컬럼 자동 추출
# =========================
date_cols = [col for col in df.columns if col.startswith("2025-")]

results = []

for idx, row in df.iterrows():
    # 날짜별 종가만 추출 (문자열 → 숫자 변환)
    prices = row[date_cols].astype(str).str.replace(",", "").str.strip()
    prices = pd.to_numeric(prices, errors="coerce").dropna()

    if len(prices) == 0:
        continue
    
    # 정규화된 시계열 생성 (첫날=100 기준)
    normalized_prices = (prices / prices.iloc[0] * 100).round(2)
    
    # 후반부 저점 패턴 감지 (반등률 조건 제거)
    is_late_trough = detect_late_trough(
        normalized_prices,
        trough_threshold=TROUGH_THRESHOLD,
        late_trough_cutoff=LATE_TROUGH_CUTOFF
    )
    
    # 모나리자 종목 디버그 (조건 만족 여부와 관계없이)
    if row["종목"] == "모나리자":
        argmin_idx = normalized_prices.values.argmin()
        trough_position = argmin_idx / (len(normalized_prices) - 1)
        drop_rate = (prices.min() / prices.iloc[0] - 1) * 100
        
        print(f"\n🔍 모나리자 디버그:")
        print(f"  - 전체 기간: {len(normalized_prices)}일")
        print(f"  - 저점 인덱스: {argmin_idx}")
        print(f"  - 저점 위치: {trough_position:.3f} ({trough_position*100:.1f}%)")
        print(f"  - 하락률: {drop_rate:.2f}%")
        print(f"  - 80% 조건 만족: {trough_position >= 0.8}")
        print(f"  - 하락률 조건 만족: {drop_rate <= -5}")
        print(f"  - 최종 조건 만족: {is_late_trough}")
    
    if is_late_trough:
        last_close = prices.iloc[-1]
        real_min = prices.min()
        
        # 정규화된 가격에서 저점 위치 계산 (detect_late_trough와 동일한 방식)
        argmin_idx = normalized_prices.values.argmin()  # 정규화된 가격의 인덱스 사용
        
        # 저점 위치 계산 (0~1 비율)
        trough_position = argmin_idx / (len(normalized_prices) - 1)
        
        # 시작 대비 저점 하락률 계산
        drop_rate = (real_min / prices.iloc[0] - 1) * 100
        
        # 디버그 정보 출력 (특정 종목 확인용)
        if row["종목"] in ["삼일제약", "모나리자"]:
            print(f"\n🔍 디버그 - {row['종목']}:")
            print(f"  - 전체 기간: {len(normalized_prices)}일")
            print(f"  - 저점 인덱스: {argmin_idx}")
            print(f"  - 저점 위치: {trough_position:.3f} ({trough_position*100:.1f}%)")
            print(f"  - 정규화된 가격 범위: {normalized_prices.min():.2f} ~ {normalized_prices.max():.2f}")
            print(f"  - 원본 가격 범위: {prices.min():.0f} ~ {prices.max():.0f}")
            print(f"  - 하락률: {drop_rate:.2f}%")
            print(f"  - 80% 조건 만족: {trough_position >= 0.8}")
            print(f"  - 하락률 조건 만족: {drop_rate <= -5}")
        
        results.append({
            "종목": row["종목"],
            "최근종가": int(last_close),
            "최저가": int(real_min),
            "저점위치(%)": round(trough_position * 100, 1),
            "하락률(%)": round(drop_rate, 2),
            "시가총액": row["시가총액"],
            "섹터": row["섹터"]
        })

# =========================
# 결과 출력 및 저장
# =========================
result_df = pd.DataFrame(results)
print("🔹 후반부 저점 패턴 종목 (80% 이후 저점, -5% 이상 하락)")
print(f"📊 총 {len(result_df)}개 종목 발견")
print(result_df.sort_values("저점위치(%)", ascending=False))

# =========================
# Excel 파일에 스파크라인 생성
# =========================
if len(result_df) > 0:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(os.path.dirname(file_path), f"후반부저점종목_스파크라인_{timestamp}.xlsx")
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "후반부저점종목"
    
    # 헤더 추가
    headers = ['종목', '스파크라인', '최근종가', '최저가', '저점위치(%)', '하락률(%)', '시가총액', '섹터']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 20  # 종목명
    ws.column_dimensions['B'].width = 20  # 스파크라인
    ws.column_dimensions['C'].width = 12  # 최근종가
    ws.column_dimensions['D'].width = 12  # 최저가
    ws.column_dimensions['E'].width = 12  # 저점위치
    ws.column_dimensions['F'].width = 12  # 하락률
    ws.column_dimensions['G'].width = 15  # 시가총액
    ws.column_dimensions['H'].width = 15  # 섹터
    
    # 행 높이 설정
    ws.row_dimensions[1].height = 30  # 헤더 행 높이
    for row in range(2, len(result_df) + 2):
        ws.row_dimensions[row].height = 50  # 스파크라인을 위한 충분한 높이
    
    # 각 종목 데이터 추가
    print(f"\n📊 Excel 스파크라인 생성 중... ({len(result_df)}개 종목)")
    temp_files = []  # 임시 파일 추적용
    
    for row_idx, (_, result_row) in enumerate(result_df.iterrows(), 2):
        try:
            # 종목명
            ws.cell(row=row_idx, column=1, value=result_row['종목'])
            
            # 원본 데이터에서 해당 종목의 가격 데이터 찾기
            original_row = df[df['종목'] == result_row['종목']].iloc[0]
            price_data = original_row[date_cols].astype(str).str.replace(",", "").str.strip()
            price_data = pd.to_numeric(price_data, errors="coerce").dropna()
            
            # 스파크라인 이미지 생성 및 삽입
            if len(price_data) > 1:
                temp_img_path = create_sparkline_image(price_data, 180, 40)
                if temp_img_path:
                    temp_files.append(temp_img_path)
                    
                    # Excel에 이미지 삽입
                    img = Image(temp_img_path)
                    img.width = 180
                    img.height = 40
                    ws.add_image(img, f'B{row_idx}')
            
            # 데이터 추가
            data_values = [
                result_row['최근종가'],
                result_row['최저가'],
                result_row['저점위치(%)'],
                result_row['하락률(%)'],
                result_row['시가총액'],
                result_row['섹터']
            ]
            
            for col_idx, value in enumerate(data_values, 3):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
            
            if row_idx % 10 == 0:
                print(f"진행률: {row_idx-1}/{len(result_df)} 종목 완료")
                
        except Exception as e:
            print(f"종목 '{result_row['종목']}' 처리 중 오류: {e}")
            continue
    
    # Excel 파일 저장
    wb.save(excel_path)
    
    # 임시 파일들 삭제
    for temp_file in temp_files:
        try:
            os.unlink(temp_file)
        except:
            pass
    
    print(f"✅ Excel 파일이 '{excel_path}'에 저장되었습니다.")
    
    # =========================
    # 저장된 파일 자동 열기
    # =========================
    if platform.system() == "Windows":
        os.startfile(excel_path)
    elif platform.system() == "Darwin":  # macOS
        os.system(f"open '{excel_path}'")
    else:  # Linux
        os.system(f"xdg-open '{excel_path}'")
    
    print("✅ Excel 파일이 열렸습니다! B열에 스파크라인을 확인하세요.")
else:
    print("❌ 조건을 만족하는 종목이 없습니다.")
