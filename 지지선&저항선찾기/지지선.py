# -*- coding: utf-8 -*-
"""
강한 지지선 분석 + 스파크라인 Excel 저장 (최근 종가 기준 ±3%)
"""

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import os
import tempfile
import warnings
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
warnings.filterwarnings('ignore')

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False


class StrongSupportWithSparkline:
    def __init__(self, csv_file_path, tolerance_pct=0.03, sideways_days=3, lookback_days=120):
        self.df = pd.read_csv(csv_file_path)

        # 시가총액, 섹터 제거
        for col in ['시가총액', '섹터']:
            if col in self.df.columns:
                self.df.drop(columns=col, inplace=True)

        if '종목' not in self.df.columns:
            raise ValueError("CSV 파일에 '종목' 컬럼이 필요합니다.")

        self.df.set_index('종목', inplace=True)
        self.tolerance_pct = tolerance_pct
        self.sideways_days = sideways_days
        self.lookback_days = lookback_days
        self.results = []

    def find_strong_support(self, price_series):
        """최근 종가 기준 ±허용범위 내 횡보 일수 확인"""
        price_series = price_series.tail(self.lookback_days)
        if len(price_series) < self.sideways_days:
            return None

        current_price = price_series.iloc[-1]  # ✅ 최근 종가 기준
        lowest_idx = price_series.idxmin()
        lowest_price = price_series.loc[lowest_idx]

        # ✅ 최근 종가 ±3% 범위 계산
        lower_bound = current_price * (1 - self.tolerance_pct)
        upper_bound = current_price * (1 + self.tolerance_pct)

        after = price_series.loc[lowest_idx:]
        in_range = (after >= lower_bound) & (after <= upper_bound)
        sideways_count = in_range.sum()

        return {
            "lowest_price": float(lowest_price),
            "current_price": float(current_price),
            "lower_bound": float(lower_bound),
            "upper_bound": float(upper_bound),
            "sideways_count": int(sideways_count),
            "is_strong_support": sideways_count >= self.sideways_days,
            "price_series": price_series
        }

    def analyze_all_stocks(self):
        print("🔍 강한 지지선 분석 시작...")
        results = []
        for stock_name, row in self.df.iterrows():
            price_data = pd.to_numeric(row, errors='coerce').dropna()
            result = self.find_strong_support(price_data)
            if result and result["is_strong_support"]:
                results.append({"stock": stock_name, **result})
        self.results = results
        print(f"📈 강한 지지선 종목: {len(results)}개 발견")

    def create_sparkline_image(self, data, width=180, height=40, color='#2E86AB'):
        data = pd.to_numeric(data, errors='coerce').dropna()
        if len(data) < 2:
            return None

        dpi = 100
        fig, ax = plt.subplots(figsize=(width/dpi, height/dpi), dpi=dpi)
        ax.plot(data.values, color=color, linewidth=2, alpha=0.8)

        max_idx = data.values.argmax()
        min_idx = data.values.argmin()
        ax.scatter([max_idx], [data.values[max_idx]], color='red', s=15)
        ax.scatter([min_idx], [data.values[min_idx]], color='blue', s=15)

        ax.set_xticks([]); ax.set_yticks([])
        for spine in ax.spines.values():
            spine.set_visible(False)
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        fig.savefig(temp_file.name, dpi=dpi, bbox_inches='tight', pad_inches=0, facecolor='white')
        plt.close(fig)
        return temp_file.name

    def save_results_to_excel(self, output_file=None):
        if not self.results:
            print("저장할 결과가 없습니다.")
            return

        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"강한지지선_with_sparklines_{timestamp}.xlsx"

        output_dir = "strong_support"
        os.makedirs(output_dir, exist_ok=True)
        filepath = os.path.join(output_dir, output_file)

        wb = Workbook()
        ws = wb.active
        ws.title = "강한 지지선 결과"

        headers = ["종목", "스파크라인", "최근종가", "최저점", "허용하단", "허용상단", "횡보일수"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        for row_idx in range(2, len(self.results) + 2):
            ws.row_dimensions[row_idx].height = 50

        temp_files = []
        for row_idx, result in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=result["stock"])
            img_path = self.create_sparkline_image(result["price_series"])
            if img_path:
                img = Image(img_path)
                img.width = 180; img.height = 40
                ws.add_image(img, f'B{row_idx}')
                temp_files.append(img_path)

            ws.cell(row=row_idx, column=3, value=result["current_price"])
            ws.cell(row=row_idx, column=4, value=result["lowest_price"])
            ws.cell(row=row_idx, column=5, value=result["lower_bound"])
            ws.cell(row=row_idx, column=6, value=result["upper_bound"])
            ws.cell(row=row_idx, column=7, value=result["sideways_count"])

        wb.save(filepath)

        for f in temp_files:
            try:
                os.unlink(f)
            except:
                pass

        print(f"✅ 스파크라인 포함 Excel 저장 완료: {filepath}")
        return filepath


def select_csv_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="CSV 파일 선택",
        filetypes=[("CSV 파일", "*.csv")]
    )
    root.destroy()
    return file_path


def main():
    print("🚀 강한 지지선 + 스파크라인 분석 시작 (최근 종가 ±3%)")
    csv_file = select_csv_file()
    if not csv_file:
        print("❌ 파일이 선택되지 않았습니다.")
        return

    analyzer = StrongSupportWithSparkline(csv_file, tolerance_pct=0.03, sideways_days=3, lookback_days=120)
    analyzer.analyze_all_stocks()
    analyzer.save_results_to_excel()


if __name__ == "__main__":
    main()
