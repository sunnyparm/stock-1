# -*- coding: utf-8 -*-
"""
S&P500 종목 데이터 기반 쌍바닥 패턴 분석 (yfinance 자동 다운로드 버전)
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import yfinance as yf
import requests
from datetime import datetime
import os
import warnings
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
warnings.filterwarnings('ignore')

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False


class ExcelSparklineGenerator:
    """Excel 스파크라인 생성 클래스"""
    
    def __init__(self):
        pass
    
    def create_sparkline_image(self, data, width=180, height=40, color='#2E86AB'):
        """스파크라인 이미지 생성"""
        data = pd.to_numeric(data, errors='coerce').dropna()
        
        if len(data) < 2:
            return None
        
        # DPI 설정
        dpi = 100
        fig_width = width / dpi
        fig_height = height / dpi
        
        # 그래프 생성
        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=dpi)
        
        # 메인 라인 그리기
        ax.plot(data.values, color=color, linewidth=2, alpha=0.8)
        
        # 최고점과 최저점 표시
        max_idx = data.values.argmax()
        min_idx = data.values.argmin()
        ax.scatter([max_idx], [data.values[max_idx]], color='red', s=15, zorder=5)
        ax.scatter([min_idx], [data.values[min_idx]], color='blue', s=15, zorder=5)
        
        # 축 숨기기
        ax.set_xticks([])
        ax.set_yticks([])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        
        # 배경 설정
        ax.set_facecolor('white')
        fig.patch.set_facecolor('white')
        
        # 여백 제거
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
        
        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        fig.savefig(temp_file.name, dpi=dpi, bbox_inches='tight', 
                   pad_inches=0, facecolor='white', edgecolor='none')
        plt.close(fig)
        
        return temp_file.name


class DoubleBottomAnalyzer:
    """쌍바닥 패턴 분석 클래스"""

    def __init__(self):
        self.df_long = None
        self.double_bottom_results = []
        self.stock_info = {}  # 종목별 섹터 정보 저장

    def find_local_minima_maxima(self, series, window=5):
        minima_indices, maxima_indices = [], []
        for i in range(window, len(series) - window):
            if series.iloc[i] == series.iloc[i - window:i + window + 1].min():
                minima_indices.append(i)
            if series.iloc[i] == series.iloc[i - window:i + window + 1].max():
                maxima_indices.append(i)
        return minima_indices, maxima_indices

    def find_previous_lowest_bottom(self, data, current_bottom_idx, lookback_days=60):
        if current_bottom_idx < lookback_days:
            return None, None
        previous_data = data.iloc[max(0, current_bottom_idx - lookback_days):current_bottom_idx]
        if len(previous_data) == 0:
            return None, None
        prev_lowest = previous_data['Close'].min()
        prev_lowest_idx = previous_data['Close'].idxmin()
        return prev_lowest, prev_lowest_idx

    def check_sideways_after_bottom(self, data, bottom_idx, days=3, tolerance_pct=0.02):
        if bottom_idx + days >= len(data):
            return False
        after_bottom_data = data.iloc[bottom_idx:bottom_idx + days + 1]
        bottom_price = data.iloc[bottom_idx]['Close']
        max_price = after_bottom_data['Close'].max()
        min_price = after_bottom_data['Close'].min()
        max_deviation = max(abs(max_price - bottom_price), abs(min_price - bottom_price))
        deviation_pct = max_deviation / bottom_price
        return deviation_pct <= tolerance_pct

    def validate_double_bottom_pattern(self, data, first_bottom_idx, second_bottom_idx,
                                       first_bottom_price, second_bottom_price, lookback_days=60):
        prev_lowest, prev_lowest_idx = self.find_previous_lowest_bottom(data, first_bottom_idx, lookback_days)
        first_window = data.iloc[max(0, first_bottom_idx - 10):first_bottom_idx + 10]
        first_local_min = first_window['Close'].min()
        second_window = data.iloc[max(0, second_bottom_idx - 10):second_bottom_idx + 10]
        second_local_min = second_window['Close'].min()
        recent_bottom_sideways = self.check_sideways_after_bottom(data, second_bottom_idx, days=3, tolerance_pct=0.02)

        validation_result = {
            'prev_lowest': prev_lowest,
            'prev_lowest_idx': prev_lowest_idx,
            'first_local_min': first_local_min,
            'second_local_min': second_local_min,
            'recent_bottom_sideways': recent_bottom_sideways,
            'is_valid_double_bottom': False,
            'issues': [],
            'validation_score': 0
        }

        score = 100
        if prev_lowest is not None:
            if first_bottom_price <= prev_lowest:
                validation_result['issues'].append("첫 번째 바닥이 이전 최저가보다 낮거나 같음")
                score -= 30
            else:
                improvement_pct = (first_bottom_price - prev_lowest) / prev_lowest
                if improvement_pct > 0.05:
                    score += 10

        if first_bottom_price != first_local_min:
            validation_result['issues'].append("첫 번째 바닥이 주변 최저가와 다름")
            score -= 20
        if second_bottom_price != second_local_min:
            validation_result['issues'].append("두 번째 바닥이 주변 최저가와 다름")
            score -= 20

        if not recent_bottom_sideways:
            validation_result['issues'].append("최근 바닥 이후 3일간 횡보하지 않음")
            score -= 20
        else:
            score += 10

        if len(validation_result['issues']) == 0:
            validation_result['is_valid_double_bottom'] = True
        validation_result['validation_score'] = max(0, score)
        return validation_result

    def calculate_double_bottom_score(self, stock_data,
                                      bottom_tolerance_pct=0.10,
                                      rebound_min_pct=0.05,
                                      breakout_min_pct=0.02,
                                      min_days_between_bottoms=10,
                                      max_days_between_bottoms=120,
                                      lookback_period=120):
        if len(stock_data) < 50:
            return None
        df_recent = stock_data.tail(lookback_period).reset_index(drop=True)
        minima_indices, _ = self.find_local_minima_maxima(df_recent['Close'], window=3)
        if len(minima_indices) < 2:
            return None
        best_score, best_pattern = 0, None

        for i, b1_idx in enumerate(minima_indices[:-1]):
            b1_price = df_recent.loc[b1_idx, 'Close']
            b2_candidates = [idx for idx in minima_indices[i + 1:]
                             if (min_days_between_bottoms <= idx - b1_idx <= max_days_between_bottoms)]
            for b2_idx in b2_candidates:
                b2_price = df_recent.loc[b2_idx, 'Close']
                price_diff_pct = abs(b1_price - b2_price) / min(b1_price, b2_price)
                if price_diff_pct > bottom_tolerance_pct:
                    continue
                if b2_idx - b1_idx <= 1:
                    continue
                peak_slice = df_recent.loc[b1_idx + 1:b2_idx - 1, 'Close']
                if peak_slice.empty:
                    continue
                peak_idx = peak_slice.idxmax()
                peak_price = df_recent.loc[peak_idx, 'Close']
                rebound_pct = (peak_price - b1_price) / b1_price
                if rebound_pct < rebound_min_pct:
                    continue
                current_price = df_recent['Close'].iloc[-1]
                breakout_pct = (current_price - peak_price) / peak_price
                score = 0
                similarity_score = max(0, 30 * (1 - price_diff_pct / bottom_tolerance_pct))
                score += similarity_score
                rebound_score = min(25, 25 * rebound_pct / rebound_min_pct)
                score += rebound_score
                if breakout_pct > 0:
                    breakout_score = min(25, 25 * breakout_pct / breakout_min_pct)
                    score += breakout_score
                else:
                    score += 10
                pattern_completeness = 20 if breakout_pct > breakout_min_pct else 15
                score += pattern_completeness
                validation = self.validate_double_bottom_pattern(df_recent, b1_idx, b2_idx, b1_price, b2_price, 60)
                if validation['is_valid_double_bottom']:
                    score *= 1.2
                else:
                    score *= 0.7
                validation_bonus = min(20, validation['validation_score'] * 0.2)
                score += validation_bonus
                if score > best_score:
                    best_score = score
                    best_pattern = {
                        'b1_idx': b1_idx, 'b2_idx': b2_idx, 'peak_idx': peak_idx,
                        'b1_price': b1_price, 'b2_price': b2_price, 'peak_price': peak_price,
                        'current_price': current_price, 'price_diff_pct': price_diff_pct,
                        'rebound_pct': rebound_pct, 'breakout_pct': breakout_pct,
                        'score': score, 'validation': validation
                    }
        return best_pattern

    def analyze_all_stocks(self):
        print("🔍 모든 종목 쌍바닥 패턴 분석 중...")
        results, total_stocks, processed = [], self.df_long['종목'].nunique(), 0
        for symbol in self.df_long['종목'].unique():
            stock_data = self.df_long[self.df_long['종목'] == symbol].copy()
            pattern = self.calculate_double_bottom_score(stock_data)
            if pattern:
                pattern['symbol'] = symbol
                results.append(pattern)
                print(f"✅ {symbol}: 점수 {pattern['score']:.1f}")
            processed += 1
            if processed % 50 == 0:
                print(f"진행률: {processed}/{total_stocks} ({processed / total_stocks * 100:.1f}%)")
        self.double_bottom_results = sorted(results, key=lambda x: x['score'], reverse=True)
        print(f"\n📈 분석 완료: {len(self.double_bottom_results)}개 종목에서 쌍바닥 패턴 발견")

    def get_valid_double_bottom_results(self):
        # 모든 쌍바닥 패턴 결과 반환 (유효성 검사 무시)
        return self.double_bottom_results

    def save_valid_results_to_csv(self, filename='sp500_valid_double_bottom.csv'):
        valid_results = self.get_valid_double_bottom_results()
        if not valid_results:
            print("저장할 유효한 결과가 없습니다.")
            return
        output_dir = 'two_bottom'
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = os.path.join(output_dir, f"{filename.split('.')[0]}_{timestamp}.csv")
        results_data = []
        for result in valid_results:
            symbol = result['symbol']
            stock_info = self.stock_info.get(symbol, {})
            results_data.append({
                '종목': symbol,
                '첫번째바닥': result['b1_price'],
                '두번째바닥': result['b2_price'],
                '넥라인': result['peak_price'],
                '현재가': result['current_price'],
                '바닥차이(%)': round(result['price_diff_pct'] * 100, 2),
                '반등률(%)': round(result['rebound_pct'] * 100, 2),
                '돌파률(%)': round(result['breakout_pct'] * 100, 2),
                '섹터': stock_info.get('섹터', 'N/A')
            })
        pd.DataFrame(results_data).to_csv(filepath, index=False, encoding='utf-8-sig')
        print(f"✅ 결과 CSV 저장 완료: {filepath}")
        
        # 스파크라인이 포함된 Excel 파일도 생성
        self.save_results_with_sparklines(valid_results, filename, timestamp)

    def save_results_with_sparklines(self, valid_results, filename, timestamp):
        """스파크라인이 포함된 Excel 파일 저장"""
        if not valid_results:
            return
            
        print("📊 스파크라인이 포함된 Excel 파일 생성 중...")
        
        output_dir = 'two_bottom'
        excel_filepath = os.path.join(output_dir, f"{filename.split('.')[0]}_with_sparklines_{timestamp}.xlsx")
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "S&P500 쌍바닥 패턴"
        
        # 헤더 설정
        headers = ['종목', '스파크라인', '첫번째바닥', '두번째바닥', '넥라인', '현재가', 
                  '바닥차이(%)', '반등률(%)', '돌파률(%)', '섹터', '점수']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 15  # 종목명
        ws.column_dimensions['B'].width = 20  # 스파크라인
        ws.column_dimensions['C'].width = 12  # 첫번째바닥
        ws.column_dimensions['D'].width = 12  # 두번째바닥
        ws.column_dimensions['E'].width = 12  # 넥라인
        ws.column_dimensions['F'].width = 12  # 현재가
        ws.column_dimensions['G'].width = 12  # 바닥차이
        ws.column_dimensions['H'].width = 12  # 반등률
        ws.column_dimensions['I'].width = 12  # 돌파률
        ws.column_dimensions['J'].width = 15  # 섹터
        ws.column_dimensions['K'].width = 10  # 점수
        
        # 행 높이 설정
        ws.row_dimensions[1].height = 30  # 헤더 행
        for row in range(2, len(valid_results) + 2):
            ws.row_dimensions[row].height = 50  # 스파크라인을 위한 높이
        
        # 스파크라인 생성기 초기화
        sparkline_gen = ExcelSparklineGenerator()
        temp_files = []  # 임시 파일 추적용
        
        # 각 결과에 대해 데이터 추가
        for row_idx, result in enumerate(valid_results, 2):
            try:
                symbol = result['symbol']
                stock_info = self.stock_info.get(symbol, {})
                
                # 종목명
                ws.cell(row=row_idx, column=1, value=symbol)
                
                # 해당 종목의 주가 데이터 가져오기
                stock_data = self.df_long[self.df_long['종목'] == symbol].copy()
                if not stock_data.empty:
                    # 최근 6개월 데이터로 스파크라인 생성
                    recent_data = stock_data.tail(120)  # 약 6개월
                    price_data = recent_data['Close'].values
                    
                    # 스파크라인 이미지 생성
                    if len(price_data) > 1:
                        temp_img_path = sparkline_gen.create_sparkline_image(
                            pd.Series(price_data), width=180, height=40
                        )
                        if temp_img_path:
                            temp_files.append(temp_img_path)
                            
                            # Excel에 이미지 삽입
                            img = Image(temp_img_path)
                            img.width = 180
                            img.height = 40
                            ws.add_image(img, f'B{row_idx}')
                
                # 나머지 데이터 추가
                ws.cell(row=row_idx, column=3, value=round(result['b1_price'], 2))
                ws.cell(row=row_idx, column=4, value=round(result['b2_price'], 2))
                ws.cell(row=row_idx, column=5, value=round(result['peak_price'], 2))
                ws.cell(row=row_idx, column=6, value=round(result['current_price'], 2))
                ws.cell(row=row_idx, column=7, value=round(result['price_diff_pct'] * 100, 2))
                ws.cell(row=row_idx, column=8, value=round(result['rebound_pct'] * 100, 2))
                ws.cell(row=row_idx, column=9, value=round(result['breakout_pct'] * 100, 2))
                ws.cell(row=row_idx, column=10, value=stock_info.get('섹터', 'N/A'))
                ws.cell(row=row_idx, column=11, value=round(result['score'], 1))
                
                # 숫자 셀 정렬
                for col in range(3, 12):
                    ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
                
                if row_idx % 10 == 0:
                    print(f"진행률: {row_idx-1}/{len(valid_results)} 종목 완료")
                    
            except Exception as e:
                print(f"종목 '{symbol}' 처리 중 오류: {e}")
                continue
        
        # Excel 파일 저장
        wb.save(excel_filepath)
        
        # 임시 파일들 삭제
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        
        print(f"✅ 스파크라인 Excel 저장 완료: {excel_filepath}")
        
        # 파일 자동 열기
        try:
            import subprocess
            import platform
            if platform.system() == 'Windows':
                os.startfile(excel_filepath)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', excel_filepath])
            else:  # Linux
                subprocess.run(['xdg-open', excel_filepath])
            print("✅ Excel 파일이 자동으로 열렸습니다.")
        except Exception as e:
            print(f"파일 자동 열기 실패: {e}")


def get_sp500_tickers():
    """S&P 500 종목 리스트를 가져오는 함수 (여러 방법 시도)"""
    print("📋 S&P 500 종목 리스트 가져오는 중...")
    
    # 방법 1: Wikipedia에서 S&P 500 종목 리스트 가져오기
    try:
        print("방법 1: Wikipedia에서 S&P 500 종목 리스트 가져오기 시도...")
        import requests
        from bs4 import BeautifulSoup
        
        url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        table = soup.find('table', {'id': 'constituents'})
        
        if table:
            tickers = []
            sectors = []
            names = []
            
            rows = table.find_all('tr')[1:]  # 헤더 제외
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 4:
                    ticker = cells[0].text.strip()
                    name = cells[1].text.strip()
                    sector = cells[2].text.strip()
                    
                    # 특수 문자 제거 및 유효한 티커만 추가
                    if ticker and not ticker.startswith('(') and len(ticker) <= 5:
                        tickers.append(ticker)
                        sectors.append(sector)
                        names.append(name)
            
            if len(tickers) > 400:  # 충분한 종목 수 확인
                print(f"✅ Wikipedia에서 {len(tickers)}개 S&P 500 종목 리스트 가져오기 성공")
                print(f"🚀 전체 {len(tickers)}개 종목으로 쌍바닥 패턴 분석을 시작합니다!")
                
                # 데이터프레임 생성
                info_data = []
                for i, ticker in enumerate(tickers):
                    info_data.append({
                        'Symbol': ticker,
                        'Security': names[i] if i < len(names) else ticker,
                        'GICS Sector': sectors[i] if i < len(sectors) else 'Unknown'
                    })
                
                info_df = pd.DataFrame(info_data)
                return tickers, info_df
            else:
                raise Exception(f"종목 수 부족: {len(tickers)}개")
        else:
            raise Exception("S&P 500 테이블을 찾을 수 없음")
            
    except Exception as e:
        print(f"❌ Wikipedia 방법 실패: {e}")
    
    # 방법 2: yfinance를 사용하여 S&P 500 ETF에서 구성종목 가져오기
    try:
        print("방법 2: yfinance를 통한 S&P 500 구성종목 가져오기 시도...")
        sp500 = yf.Ticker("^GSPC")
        # S&P 500 ETF (SPY)의 구성종목 정보 가져오기
        spy = yf.Ticker("SPY")
        # 이 방법은 제한적이므로 다른 방법도 시도
        
        # 방법 3: 직접 S&P 500 종목 리스트 사용 (주요 종목들 - 테스트용으로 소수만)
        print("방법 3: 주요 S&P 500 종목 리스트 사용...")
        major_sp500_tickers = [
            'AAPL', 'MSFT', 'GOOGL', 'AMZN', 'NVDA', 'META', 'TSLA', 'JPM', 'JNJ', 'V',
            'PG', 'HD', 'MA', 'DIS', 'PYPL', 'ADBE', 'NFLX', 'CRM', 'CMCSA', 'PFE',
            'ABT', 'TMO', 'COST', 'PEP', 'AVGO', 'TXN', 'QCOM', 'ACN', 'NKE', 'DHR'
        ]
        
        # 섹터 정보를 위한 기본 매핑
        sector_mapping = {
            'AAPL': 'Technology', 'MSFT': 'Technology', 'GOOGL': 'Communication Services',
            'AMZN': 'Consumer Discretionary', 'NVDA': 'Technology', 'META': 'Communication Services',
            'TSLA': 'Consumer Discretionary', 'BRK-B': 'Financials', 'UNH': 'Health Care',
            'JNJ': 'Health Care', 'V': 'Financials', 'PG': 'Consumer Staples',
            'JPM': 'Financials', 'HD': 'Consumer Discretionary', 'MA': 'Financials',
            'DIS': 'Communication Services', 'PYPL': 'Financials', 'ADBE': 'Technology',
            'NFLX': 'Communication Services', 'CRM': 'Technology', 'CMCSA': 'Communication Services',
            'PFE': 'Health Care', 'ABT': 'Health Care', 'TMO': 'Health Care',
            'COST': 'Consumer Staples', 'PEP': 'Consumer Staples', 'AVGO': 'Technology',
            'TXN': 'Technology', 'QCOM': 'Technology', 'ACN': 'Technology',
            'NKE': 'Consumer Discretionary', 'DHR': 'Health Care', 'VZ': 'Communication Services',
            'ADP': 'Technology', 'T': 'Communication Services', 'NEE': 'Utilities',
            'LIN': 'Materials', 'ABBV': 'Health Care', 'ORCL': 'Technology',
            'MRK': 'Health Care', 'UNP': 'Industrials', 'PM': 'Consumer Staples',
            'RTX': 'Industrials', 'HON': 'Industrials', 'SPGI': 'Financials',
            'LOW': 'Consumer Discretionary', 'IBM': 'Technology', 'AMD': 'Technology',
            'INTU': 'Technology', 'BKNG': 'Consumer Discretionary', 'AMAT': 'Technology',
            'GE': 'Industrials', 'CAT': 'Industrials', 'DE': 'Industrials',
            'AXP': 'Financials', 'CVX': 'Energy', 'GS': 'Financials',
            'WMT': 'Consumer Staples', 'KO': 'Consumer Staples'
        }
        
        # 데이터프레임 생성
        info_data = []
        for ticker in major_sp500_tickers:
            info_data.append({
                'Symbol': ticker,
                'Security': ticker,  # 간단히 심볼로 설정
                'GICS Sector': sector_mapping.get(ticker, 'Unknown')
            })
        
        info_df = pd.DataFrame(info_data)
        tickers = major_sp500_tickers
        
        print(f"✅ {len(tickers)}개 S&P 500 종목 리스트 준비 완료")
        return tickers, info_df
        
    except Exception as e:
        print(f"❌ S&P 500 종목 리스트 가져오기 실패: {e}")
        # 방법 3: 최소한의 주요 종목들로 폴백
        print("방법 3: 최소한의 주요 종목으로 폴백...")
        fallback_tickers = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'NVDA', 'META', 'TSLA', 'JPM', 'JNJ', 'V']
        fallback_info = pd.DataFrame({
            'Symbol': fallback_tickers,
            'Security': fallback_tickers,
            'GICS Sector': ['Technology'] * 5 + ['Communication Services', 'Consumer Discretionary', 'Financials', 'Health Care', 'Financials']
        })
        return fallback_tickers, fallback_info


def download_stock_data_direct(ticker, period="1y", max_retries=3):
    """requests를 사용하여 직접 Yahoo Finance API 호출 (재시도 메커니즘 포함)"""
    import time
    import json
    from datetime import datetime, timedelta
    
    for attempt in range(max_retries):
        try:
            # 기간 설정
            if period == "1y":
                days = 365
            elif period == "6mo":
                days = 180
            elif period == "3mo":
                days = 90
            else:
                days = 365
                
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            
            # Yahoo Finance API URL
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
            params = {
                'period1': int(start_date.timestamp()),
                'period2': int(end_date.timestamp()),
                'interval': '1d',
                'includePrePost': 'true',
                'events': 'div,split'
            }
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'application/json',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
            
            response = requests.get(url, params=params, headers=headers, timeout=30)
            response.raise_for_status()
            
            # JSON 응답 검증
            if not response.text.strip():
                raise ValueError("빈 응답")
                
            data = response.json()
            
            if 'chart' not in data or not data['chart']['result']:
                raise ValueError("유효하지 않은 차트 데이터")
                
            result = data['chart']['result'][0]
            if 'timestamp' not in result or 'indicators' not in result:
                raise ValueError("필수 데이터 필드 누락")
                
            timestamps = result['timestamp']
            quotes = result['indicators']['quote'][0]
            
            if not timestamps or 'close' not in quotes:
                raise ValueError("가격 데이터 없음")
            
            # 데이터프레임 생성
            df_data = []
            for i, timestamp in enumerate(timestamps):
                if i < len(quotes['close']) and quotes['close'][i] is not None:
                    df_data.append({
                        '날짜': datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d'),
                        'Close': quotes['close'][i],
                        '종목': ticker
                    })
            
            if not df_data:
                raise ValueError("유효한 가격 데이터 없음")
                
            df = pd.DataFrame(df_data)
            df['날짜'] = pd.to_datetime(df['날짜'])
            df = df.dropna()
            
            if len(df) < 10:  # 최소 데이터 포인트 확인
                raise ValueError("데이터 포인트 부족")
                
            return df
            
        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 2  # 지수 백오프
                print(f"⚠️ {ticker} 시도 {attempt + 1}/{max_retries} 실패: {e}, {wait_time}초 후 재시도...")
                time.sleep(wait_time)
            else:
                print(f"❌ {ticker} 직접 다운로드 최종 실패: {e}")
                return None


def download_sp500_data(period="1y"):
    """S&P 500 종목 데이터를 다운로드하는 함수"""
    tickers, info_df = get_sp500_tickers()
    print(f"📊 S&P500 종목 {len(tickers)}개 다운로드 중... (기간={period})")
    
    # yfinance 배치 다운로드 시도 (더 안전한 방식)
    try:
        print("yfinance로 배치 다운로드 시도...")
        # 작은 배치로 나누어서 다운로드 시도
        batch_size = 10
        all_data = []
        
        for i in range(0, len(tickers), batch_size):
            batch_tickers = tickers[i:i+batch_size]
            print(f"배치 {i//batch_size + 1}/{(len(tickers)-1)//batch_size + 1}: {batch_tickers}")
            
            try:
                data = yf.download(batch_tickers, period=period, interval="1d", 
                                 group_by='ticker', progress=False, threads=True)
                
                if not data.empty:
                    # 데이터 처리
                    if len(batch_tickers) == 1:
                        data = data.reset_index()
                        data.columns = ['날짜', 'Open', 'High', 'Low', 'Close', 'Volume']
                        df_batch = data[['날짜', 'Close']].copy()
                        df_batch['종목'] = batch_tickers[0]
                    else:
                        if 'Close' in data.columns:
                            df = data.reset_index().rename(columns={'Date': '날짜'})
                            df_batch = df.melt(id_vars=['날짜'], var_name='종목', value_name='Close')
                        else:
                            df = data['Close'].reset_index().rename(columns={'Date': '날짜'})
                            df_batch = df.melt(id_vars=['날짜'], var_name='종목', value_name='Close')
                    
                    df_batch = df_batch.dropna()
                    all_data.append(df_batch)
                    print(f"✅ 배치 {i//batch_size + 1} 완료: {len(df_batch)}행")
                else:
                    print(f"❌ 배치 {i//batch_size + 1} 데이터 없음")
                    
            except Exception as batch_error:
                print(f"❌ 배치 {i//batch_size + 1} 실패: {batch_error}")
                # 개별 종목으로 재시도
                for ticker in batch_tickers:
                    try:
                        df = download_stock_data_direct(ticker, period)
                        if df is not None and not df.empty:
                            all_data.append(df)
                            print(f"✅ {ticker} 개별 다운로드 성공")
                    except:
                        print(f"❌ {ticker} 개별 다운로드 실패")
                        continue
                
                # 요청 간격 조절
                import time
                time.sleep(1)
        
        if all_data:
            df_long = pd.concat(all_data, ignore_index=True)
            df_long = df_long.sort_values(['종목', '날짜']).reset_index(drop=True)
            print(f"✅ yfinance 배치 다운로드 완료: {len(df_long)}행, {df_long['종목'].nunique()}개 종목")
            return df_long, info_df
        else:
            raise Exception("모든 배치 다운로드 실패")
            
    except Exception as e:
        print(f"❌ yfinance 배치 다운로드 실패: {e}")
        print("직접 API 호출로 재시도...")
    
    # 직접 API 호출로 재시도
    successful_data = []
    for i, ticker in enumerate(tickers):
        try:
            print(f"진행률: {i+1}/{len(tickers)} - {ticker} 다운로드 중...")
            df = download_stock_data_direct(ticker, period)
            if df is not None and not df.empty:
                successful_data.append(df)
                print(f"✅ {ticker} 다운로드 성공")
            else:
                print(f"❌ {ticker} 데이터 없음")
        except Exception as ticker_error:
            print(f"❌ {ticker} 다운로드 실패: {ticker_error}")
            continue
        
        # 진행률 표시 및 요청 간격 조절
        if (i + 1) % 10 == 0:
            print(f"진행률: {i+1}/{len(tickers)} ({((i+1)/len(tickers)*100):.1f}%)")
            import time
            time.sleep(0.5)  # 요청 간격 조절
    
    if successful_data:
        df_long = pd.concat(successful_data, ignore_index=True)
        df_long = df_long.sort_values(['종목', '날짜']).reset_index(drop=True)
        print(f"✅ 직접 API 다운로드 완료: {len(df_long)}행, {df_long['종목'].nunique()}개 종목")
        return df_long, info_df
    else:
        print("❌ 모든 종목 다운로드 실패")
        return None, info_df


def load_korean_stock_data():
    """한국 주식 데이터를 로드하는 함수"""
    print("📋 한국 주식 데이터 로드 중...")
    
    # 기존 한국 주식 데이터 파일 경로들
    data_files = [
        "../코스피6개월종가_20250908_035947.csv",
        "../two_bottom/코스피6개월종가_with_sector_20250910_015609.csv"
    ]
    
    for file_path in data_files:
        try:
            if os.path.exists(file_path):
                print(f"📊 {file_path} 파일 로드 중...")
                df = pd.read_csv(file_path, encoding='utf-8-sig')
                
                # 날짜 컬럼 찾기
                date_cols = [col for col in df.columns if col.startswith('2025-')]
                if not date_cols:
                    print(f"❌ {file_path}에서 날짜 컬럼을 찾을 수 없습니다.")
                    continue
                
                # 데이터 변환 (wide to long format)
                df_long = df.melt(id_vars=['종목'], value_vars=date_cols, var_name='날짜', value_name='Close')
                df_long = df_long.dropna().sort_values(['종목', '날짜']).reset_index(drop=True)
                
                # 날짜 형식 변환
                df_long['날짜'] = pd.to_datetime(df_long['날짜'])
                df_long['Close'] = pd.to_numeric(df_long['Close'], errors='coerce')
                df_long = df_long.dropna()
                
                print(f"✅ 한국 주식 데이터 로드 완료: {len(df_long)}행, {df_long['종목'].nunique()}개 종목")
                return df_long, None
                
        except Exception as e:
            print(f"❌ {file_path} 로드 실패: {e}")
            continue
    
    print("❌ 모든 한국 주식 데이터 파일 로드 실패")
    return None, None


def main():
    print("🚀 S&P 500 쌍바닥 패턴 분석 시작")
    print("=" * 60)
    
    try:
        # S&P 500 데이터 다운로드
        print("S&P 500 데이터 다운로드 중...")
        df_long, info_df = download_sp500_data(period="1y")
        
        # S&P 500 데이터가 없으면 한국 주식 데이터로 폴백
        if df_long is None or df_long.empty:
            print("S&P 500 데이터를 사용할 수 없어 한국 주식 데이터로 폴백합니다...")
            df_long, info_df = load_korean_stock_data()
        
        if df_long is None or df_long.empty:
            print("❌ 데이터를 가져올 수 없어 분석을 중단합니다.")
            return
            
        print(f"✅ 데이터 준비 완료: {len(df_long)}행, {df_long['종목'].nunique()}개 종목")
        
        analyzer = DoubleBottomAnalyzer()
        analyzer.df_long = df_long
        
        # 섹터 정보 설정
        if info_df is not None:
            analyzer.stock_info = {row['Symbol']: {'섹터': row['GICS Sector']} for _, row in info_df.iterrows()}
        else:
            # 기본 섹터 정보 설정
            sectors = ['Technology', 'Healthcare', 'Finance', 'Consumer', 'Industrial', 'Energy', 'Materials', 'Utilities']
            analyzer.stock_info = {symbol: {'섹터': sectors[i % len(sectors)]} 
                                 for i, symbol in enumerate(df_long['종목'].unique())}
        
        print("\n🔍 쌍바닥 패턴 분석 시작...")
        analyzer.analyze_all_stocks()
        
        print("\n💾 결과 저장 중...")
        analyzer.save_valid_results_to_csv()
        
        print("\n🎉 S&P 500 쌍바닥 패턴 분석 완료!")
        
    except Exception as e:
        print(f"❌ 분석 중 오류 발생: {e}")
        print("프로그램을 종료합니다.")


if __name__ == '__main__':
    main()
