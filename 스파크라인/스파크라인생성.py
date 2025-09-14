import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import os
import tempfile
import warnings
import tkinter as tk
from tkinter import filedialog, messagebox
warnings.filterwarnings('ignore')

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

class ExcelSparklineGenerator:
    def __init__(self, csv_file_path):
        """Excel 스파크라인 생성기 초기화"""
        self.df = pd.read_csv(csv_file_path)
        
        # 시가총액과 섹터 열 제외
        columns_to_exclude = ['시가총액', '섹터']
        existing_exclude_cols = [col for col in columns_to_exclude if col in self.df.columns]
        
        if existing_exclude_cols:
            print(f"제외할 열: {existing_exclude_cols}")
            self.df = self.df.drop(columns=existing_exclude_cols)
        
        self.df.set_index('종목', inplace=True)
        
    def create_sparkline_image(self, data, width=180, height=40, color='#2E86AB'):
        """스파크라인 이미지 생성"""
        data = pd.to_numeric(data, errors='coerce').dropna()
        
        if len(data) < 2:
            return None
        
        # DPI 설정
        
        dpi = 100
        fig_width = width / dpi
        fig_height = height / dpi
        
        # 그래프 생성
        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=dpi)
        
        # 메인 라인 그리기
        ax.plot(data.values, color=color, linewidth=2, alpha=0.8)
        
        # 최고점과 최저점 표시
        max_idx = data.values.argmax()
        min_idx = data.values.argmin()
        ax.scatter([max_idx], [data.values[max_idx]], color='red', s=15, zorder=5)
        ax.scatter([min_idx], [data.values[min_idx]], color='blue', s=15, zorder=5)
        
        # 축 숨기기
        ax.set_xticks([])
        ax.set_yticks([])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        
        # 배경 설정
        ax.set_facecolor('white')
        fig.patch.set_facecolor('white')
        
        # 여백 제거
        plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
        
        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        fig.savefig(temp_file.name, dpi=dpi, bbox_inches='tight', 
                   pad_inches=0, facecolor='white', edgecolor='none')
        plt.close(fig)
        
        return temp_file.name
    
    def create_excel_with_sparklines(self, output_file=None, start_col=2, max_rows=None):
        """Excel 파일에 스파크라인을 삽입하여 생성"""
        if output_file is None:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"코스피6개월종가_with_sparklines_{timestamp}.xlsx"
        
        if max_rows is None:
            max_rows = len(self.df)
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "스파크라인"
        
        # 헤더 추가
        headers = ['종목', '스파크라인'] + list(self.df.columns[start_col-1:])
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 컬럼 너비 설정
        ws.column_dimensions['A'].width = 20  # 종목명
        ws.column_dimensions['B'].width = 20  # 스파크라인
        
        # 행 높이 설정 (스파크라인이 잘리지 않도록)
        ws.row_dimensions[1].height = 30  # 헤더 행 높이
        for row in range(2, max_rows + 2):  # 데이터 행들
            ws.row_dimensions[row].height = 50  # 스파크라인을 위한 충분한 높이
        
        # 각 종목 데이터 추가
        print(f"Excel 스파크라인 생성 중... ({max_rows}개 종목)")
        temp_files = []  # 임시 파일 추적용
        
        for row_idx, (stock_name, row) in enumerate(self.df.iterrows(), 2):
            if row_idx > max_rows + 1:  # max_rows + 헤더 행
                break
                
            try:
                # 종목명
                ws.cell(row=row_idx, column=1, value=stock_name)
                
                # 종가 데이터 추출
                price_data = row.iloc[start_col-1:]
                price_data = pd.to_numeric(price_data, errors='coerce').dropna()
                
                # 스파크라인 이미지 생성 및 삽입
                if len(price_data) > 1:
                    temp_img_path = self.create_sparkline_image(price_data, 180, 40)
                    if temp_img_path:
                        temp_files.append(temp_img_path)  # 나중에 삭제하기 위해 추적
                        
                        # Excel에 이미지 삽입
                        img = Image(temp_img_path)
                        img.width = 180
                        img.height = 40
                        ws.add_image(img, f'B{row_idx}')
                
                # 가격 데이터 추가
                for col_idx, value in enumerate(price_data, 3):
                    if pd.notna(value):
                        cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
                        cell.alignment = Alignment(horizontal='center')
                
                if row_idx % 50 == 0:
                    print(f"진행률: {row_idx-1}/{min(max_rows, len(self.df))} 종목 완료")
                    
            except Exception as e:
                print(f"종목 '{stock_name}' 처리 중 오류: {e}")
                continue
        
        # Excel 파일 저장
        wb.save(output_file)
        
        # 임시 파일들 삭제
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        
        print(f"Excel 파일이 '{output_file}'에 저장되었습니다.")
        return output_file

def select_csv_file():
    """CSV 파일 선택 대화상자"""
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우 숨기기
    
    file_path = filedialog.askopenfilename(
        title="CSV 파일을 선택하세요",
        filetypes=[
            ("CSV 파일", "*.csv"),
            ("모든 파일", "*.*")
        ]
    )
    
    root.destroy()
    return file_path


def validate_csv_file(file_path):
    """CSV 파일 유효성 검사"""
    try:
        # 파일 존재 확인
        if not os.path.exists(file_path):
            return False, "파일이 존재하지 않습니다."
        
        # CSV 파일 읽기 시도
        df = pd.read_csv(file_path)
        
        # '종목' 컬럼 확인
        if '종목' not in df.columns:
            return False, "CSV 파일에 '종목' 컬럼이 필요합니다."
        
        # 시가총액과 섹터 열 제외 후 컬럼 수 확인
        columns_to_exclude = ['시가총액', '섹터']
        existing_exclude_cols = [col for col in columns_to_exclude if col in df.columns]
        df_filtered = df.drop(columns=existing_exclude_cols)
        
        # 최소 컬럼 수 확인 (종목 + 최소 2개 이상의 가격 데이터)
        if len(df_filtered.columns) < 3:
            return False, "CSV 파일에 종목과 최소 2개 이상의 가격 데이터 컬럼이 필요합니다."
        
        return True, f"파일이 유효합니다. (총 {len(df.columns)}개 컬럼, 가격 데이터 {len(df_filtered.columns)-1}개)"
        
    except Exception as e:
        return False, f"파일 읽기 오류: {str(e)}"

def main():
    """메인 실행 함수"""
    print("=== Excel 스파크라인 생성기 ===\n")
    
    # 1. CSV 파일 선택
    print("1. CSV 파일을 선택하세요...")
    csv_file = select_csv_file()
    
    if not csv_file:
        print("파일 선택이 취소되었습니다.")
        return
    
    print(f"선택된 파일: {os.path.basename(csv_file)}")
    
    # 2. 파일 유효성 검사
    print("\n2. 파일 유효성 검사 중...")
    is_valid, message = validate_csv_file(csv_file)
    
    if not is_valid:
        print(f"오류: {message}")
        return
    
    print(f"✓ {message}")
    
    try:
        # 3. 스파크라인 생성기 초기화
        print("\n3. 데이터 로딩 중...")
        sparkline_gen = ExcelSparklineGenerator(csv_file)
        
        print(f"✓ 사용 가능한 종목 수: {len(sparkline_gen.df)}")
        print(f"✓ 데이터 컬럼 수: {len(sparkline_gen.df.columns)}")
        
        # 4. Excel 파일 생성 (전체 종목 처리)
        max_rows = len(sparkline_gen.df)
        print(f"\n4. Excel 파일 생성 중... (전체 {max_rows}개 종목)")
        
        # 출력 파일명 생성 (CSV 파일과 같은 폴더에 저장)
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"스파크라인_{os.path.splitext(os.path.basename(csv_file))[0]}_{timestamp}.xlsx"
        output_path = os.path.join(os.path.dirname(csv_file), output_filename)
        
        print(f"저장 위치: {output_path}")
        
        excel_output = sparkline_gen.create_excel_with_sparklines(
            output_file=output_path, 
            max_rows=max_rows
        )
        
        print(f"\n✓ 생성 완료!")
        print(f"✓ Excel 파일: {output_filename}")
        print(f"✓ 저장 위치: {os.path.dirname(csv_file)}")
        print(f"✓ Excel 파일을 열어보시면 B열에 실제 스파크라인이 표시됩니다!")
        
        # 5. 파일 자동 열기
        print(f"\n5. Excel 파일을 자동으로 열기...")
        try:
            os.startfile(excel_output)  # Windows
            print("✓ Excel 파일이 열렸습니다!")
        except:
            print("파일을 자동으로 열 수 없습니다. 수동으로 열어주세요.")
        
    except Exception as e:
        print(f"\n오류가 발생했습니다: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
