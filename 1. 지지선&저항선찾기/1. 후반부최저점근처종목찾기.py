import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
import os
import platform
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import tempfile
import warnings
warnings.filterwarnings('ignore')

plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

def detect_late_trough(series: pd.Series, cutoff: float = 0.8) -> bool:
    s = series.dropna().reset_index(drop=True)
    if len(s) < 10:
        return False
    idx = int(np.argmin(s.values))
    return idx / (len(s) - 1) >= cutoff

def create_sparkline_image(data, width=180, height=40, color='#2E86AB'):
    data = pd.to_numeric(data, errors='coerce').dropna()
    if len(data) < 2:
        return None
    dpi = 100
    fig, ax = plt.subplots(figsize=(width/dpi, height/dpi), dpi=dpi)
    ax.plot(data.values, color=color, linewidth=2, alpha=0.8)
    ax.scatter([data.values.argmax()], [data.max()], color='red', s=15)
    ax.scatter([data.values.argmin()], [data.min()], color='blue', s=15)
    ax.axis('off')
    fig.patch.set_facecolor('white')
    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    fig.savefig(temp_file.name, dpi=dpi, bbox_inches='tight', pad_inches=0, facecolor='white')
    plt.close(fig)
    return temp_file.name

# 사용자 설정
CUTOFF = 0.8 # 전체 기간의 80% 이후에 저점이 있는지 판단

# 파일 선택
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="CSV 파일 선택", filetypes=[("CSV files", "*.csv")])
df = pd.read_csv(file_path)

# 날짜 컬럼 식별
date_cols = [col for col in df.columns if col.startswith("2025-")]
results = []

# 종목별 후반부 저점 필터링
for _, row in df.iterrows():
    prices = pd.to_numeric(
        row[date_cols].astype(str).str.replace(",", "").str.strip(), errors="coerce"
    ).dropna().reset_index(drop=True)

    if len(prices) < 10:
        continue

    norm_prices = (prices / prices.iloc[0]) * 100
    if not detect_late_trough(norm_prices, cutoff=CUTOFF):
        continue

    trough_idx = int(np.argmin(norm_prices))
    trough_pos = round(trough_idx / (len(norm_prices) - 1) * 100, 1)
    drop_rate = round((prices.min() / prices.iloc[0] - 1) * 100, 2)

    results.append({
        "종목": row["종목"].strip(),
        "최근종가": int(prices.iloc[-1]),
        "최저가": int(prices.min()),
        "저점위치(%)": trough_pos,
        "하락률(%)": drop_rate,
        "시가총액": row.get("시가총액", ""),
        "섹터": row.get("섹터", "")
    })

# 결과 DataFrame 생성
result_df = pd.DataFrame(results)
print(f"\n📊 후반부 저점 종목 수: {len(result_df)}개")
if result_df.empty:
    print("❌ 조건을 만족하는 종목이 없습니다.")
    exit()

# 엑셀 저장 준비
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_path = os.path.join(os.path.dirname(file_path), f"후반부저점종목_스파크라인_{timestamp}.xlsx")
wb = Workbook()
ws = wb.active
ws.title = "후반부저점종목"

headers = ['종목', '스파크라인', '최근종가', '최저가', '저점위치(%)', '하락률(%)', '시가총액', '섹터']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[chr(64 + col)].width = 15
ws.row_dimensions[1].height = 30

# 종목별 행 추가 및 스파크라인 삽입
temp_files = []
for row_idx, (_, row) in enumerate(result_df.iterrows(), start=2):
    ws.cell(row=row_idx, column=1, value=row['종목'])
    ws.row_dimensions[row_idx].height = 30

    original = df[df["종목"].astype(str).str.strip() == row['종목']].iloc[0]
    price_data = pd.to_numeric(
        original[date_cols].astype(str).str.replace(",", "").str.strip(),
        errors="coerce"
    ).dropna()

    img_path = create_sparkline_image(price_data)
    if img_path:
        temp_files.append(img_path)
        img = Image(img_path)
        img.width, img.height = 180, 40
        ws.add_image(img, f'B{row_idx}')

    values = [
        row['최근종가'],
        row['최저가'],
        row['저점위치(%)'],
        row['하락률(%)'],
        row['시가총액'],
        row['섹터']
    ]
    for col_idx, val in enumerate(values, 3):
        ws.cell(row=row_idx, column=col_idx, value=val).alignment = Alignment(horizontal='center')

# 저장 및 임시파일 정리
wb.save(excel_path)
for f in temp_files:
    try:
        os.unlink(f)
    except:
        pass

# 자동 열기
print(f"\n✅ Excel 저장 완료: {excel_path}")
if platform.system() == "Windows":
    os.startfile(excel_path)
elif platform.system() == "Darwin":
    os.system(f"open '{excel_path}'")
else:
    os.system(f"xdg-open '{excel_path}'")
